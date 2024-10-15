
#run command : pytest -s -k test_addstudent_multiple_data.py  Pytest_Project_API


import requests
import json
import openpyxl

def test_addStudent_multiple_data():
    url = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("Pytest_Project_API/Data_Driven_testing_Pytest/student_details.json", 'r')
    json_request = json.loads(f.read())
    wk = openpyxl.load_workbook('Pytest_Project_API/Data_Driven_testing_Pytest/multiple_student_details.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row

    for i in range(2, rows+1):
        first_name = sh.cell(row=i,column=1)
        middle_name = sh.cell(row=i, column=2)
        last_name = sh.cell(row=i, column=3)
        Date_of_birth = sh.cell(row=i, column=4)
        json_request['first_name'] = first_name
        json_request['middle_name'] = middle_name
        json_request['last_name'] = last_name
        json_request['date_of_birth'] = Date_of_birth
        response = requests.post(url, json_request)
        print(response.status_code)
        print(response.text)
        assert response.status_code == 201








